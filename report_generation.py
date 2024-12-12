from tkinter import messagebox
import openpyxl
from db_connection import execute_global_report_query, execute_client_report_query, execute_partida_report_query

def generate_global_report(window, database, selected_agent, start_date, end_date):
    """Generar el reporte global (concentrado) de ventas"""
    try:
        data = execute_global_report_query(database, selected_agent, start_date, end_date)
        if not data:
            messagebox.showwarning("No se encontraron datos", "No se encontraron datos para el reporte.")
            return
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Global"
        headers = ["Serie", "Folio", "Cliente", "Fecha Elaboración", "Costo Material", "Neto Documento", "Utilidad Final", "Margen", "Comisión", "Agente Venta"]
        ws.append(headers)
        
        for row in data:
            ws.append(list(row))

        filename = f"reporte_global_{selected_agent}_{start_date}_{end_date}.xlsx"
        wb.save(filename)
        messagebox.showinfo("Reporte Generado", f"Reporte generado exitosamente: {filename}")
    except Exception as e:
        messagebox.showerror("Error", f"Hubo un error al generar el reporte global: {e}")

def generate_client_report(window, database, selected_agent, start_date, end_date):
    """Generar el reporte por cliente"""
    try:
        data = execute_client_report_query(database, selected_agent, start_date, end_date)
        if not data:
            messagebox.showwarning("No se encontraron datos", "No se encontraron datos para el reporte.")
            return
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte por Cliente"
        headers = ["Cliente", "Costo Material", "Neto Documento", "Utilidad Final", "Margen", "Comisión", "Agente Venta"]
        ws.append(headers)
        
        for row in data:
            ws.append(list(row))

        filename = f"reporte_cliente_{selected_agent}_{start_date}_{end_date}.xlsx"
        wb.save(filename)
        messagebox.showinfo("Reporte Generado", f"Reporte generado exitosamente: {filename}")
    except Exception as e:
        messagebox.showerror("Error", f"Hubo un error al generar el reporte por cliente: {e}")

def generate_partida_report(database, selected_agent, start_date, end_date):
    """Generar el reporte por partida"""
    try:
        data = execute_partida_report_query(database, selected_agent, start_date, end_date)
        if not data:
            messagebox.showwarning("No se encontraron datos", "No se encontraron datos para el reporte.")
            return
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte por Partida"
        headers = ["Serie", "Folio", "Partida", "Fecha Elaboración", "Costo Material", "Neto Documento", "Utilidad Final", "Margen", "Comisión", "Agente Venta"]
        ws.append(headers)
        
        for row in data:
            ws.append(list(row))

        filename = f"reporte_partida_{selected_agent}_{start_date}_{end_date}.xlsx"
        wb.save(filename)
        messagebox.showinfo("Reporte Generado", f"Reporte generado exitosamente: {filename}")
    except Exception as e:
        messagebox.showerror("Error", f"Hubo un error al generar el reporte por partida: {e}")
